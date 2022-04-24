import tensorflow as tf
from tensorflow import keras
import numpy as np
import cv2
import random
import tkinter
from tkinter import *
import tkinter as tk

window=Tk()
window.title('CHATBOT FOR DIET PLAN AND HEALTH CARE')
window.geometry('1500x1500')
#window.configure(background='paleGreen3')
img = PhotoImage(file="4.png")
label = Label(
      window,
      image=img)
label.place(x=0,y=0,width=1400,height=800)
l1=Label(window,text='CHATBOT FOR DIET PLAN AND HEALTH CARE ',bg='darkslateblue',width=50,height=1,font=('times new roman',20,'bold'))
l1.place(x=250,y=0)
#Loading themodel
batch_size = 32
img_height = 64
img_width = 64
model_dl = keras.models.load_model("model.h5") #look for local saved file

def BMI(height, weight):
    bmi = weight/(height**2)
    return bmi

from tensorflow.keras.preprocessing import image

#Creating a dictionary to map each of the indexes to the corresponding number or letter

from tkinter.filedialog import askopenfile

dict = {0:'Apple pie',1:'Baby back ribs',2:'Baklava',3:'Beef carpaccio',4:'Beef tartare',5:'Beet salad',6:'Beignets',7:'Bibimbap',8:'Bread pudding',9:'Breakfast burrito',
        10:'Bruschetta',11:'Caesar salad',12:'Cannoli',13:'Caprese salad',14:'Carrot cake',15:'Ceviche',16:'Cheesecake',17:'Cheese plate',18:'Chicken curry',19:'Chicken quesadilla',
        20:'Chicken wings',21:'Chocolate cake',22:'Chocolate mousse',23:'Churros',24:'Clam chowder',25:'Club sandwich',26:'Crab cakes',27:'Creme brulee',28:'Croque madame',
        29:'Cup cakes',30:'Deviled eggs',31:'Donuts',32:'Dumplings',33:'Edamame',34:'Eggs benedict',35:'Escargots',36:'Falafel',37:'Filet mignon',38:'Fish and chips',
        39:'Foie gras',40:'French fries',41:'French onion soup',42:'French toast',43:'Fried calamari',44:'Fried rice',45:'Frozen yogurt',46:'Garlic bread',47:'Gnocchi',
        48:'Greek salad',49:'Grilled cheese sandwich',50:'Grilled salmon',51:'Guacamole',52:'Gyoza',53:'Hamburger',54:'Hot and sour soup',55:'Hot dog',56:'Huevos rancheros',
        57:'Hummus',58:'Ice cream',59:'Lasagna',60:'Lobster bisque',61:'Lobster roll sandwich',62:'Macaroni and cheese',63:'Macarons',64:'Miso soup',65:'Mussels',
        66:'Nachos',67:'Omelette',68:'Onion rings',69:'Oysters',70:'Pad thai',71:'Paella',72:'Pancakes',73:'Panna cotta',74:'Peking duck',75:'Pho',76:'Pizza',77:'Pork chop',
        78:'Poutine',79:'Prime rib',80:'Pulled pork sandwich',81:'Ramen',82:'Ravioli',83:'Red velvet cake',84:'Risotto',85:'Samosa',86:'Sashimi',87:'Scallops',88:'Seaweed salad',
        89:'Shrimp and grits',90:'Spaghetti bolognese',91:'Spaghetti carbonara',92:'Spring rolls',93:'Steak',94:'Strawberry shortcake',95:'Sushi',96:'Tacos',97:'Takoyaki',
        98:'Tiramisu',99:'Tuna tartare',100:'Waffles'}
def open():
     global text
     file = askopenfile(mode='r',filetypes = [('All files','*.*')])
     img_to_detect = cv2.imread(file.name, cv2.IMREAD_COLOR)
     img = cv2.resize(img_to_detect,(64,64))
     x = image.img_to_array(img) 
     x = np.expand_dims(x, axis=0)
     imag = np.vstack([x])
     classes = model_dl.predict_classes(imag, batch_size=batch_size)
     text = str(dict[classes.item()])
     #print('Bot: ',format(text))
     #a=Text(window, height=10, width=40,bg="White", fg="black",relief='flat',font=('Times',15,'bold'))
     #a.insert(END,text)
     #a.place(x=600,y=500)
     #print(text)

     from openpyxl.reader.excel import load_workbook
     wrkbk = load_workbook(r"C:\Users\Niranjan\chat bot health care 2\calorie.xlsx")
     sh = wrkbk.active
     for i in range(1,101):
      c=str(sh.cell(row=i,column=1).value)
      #print(c)
      if c == text:
        calo =str(sh.cell(row=i,column=2).value)
        a=Text(window, height=10, width=40,bg="White", fg="black",relief='flat',font=('Times',15,'bold'))
        a.insert(END,'Bot:Its a junk food\ncalories',text,format(calo))
        a.place(x=700,y=200)
        print(type(text))
        #print('Bot: The food contain',format(calo),'calories')
        typr =str(sh.cell(row=i,column=4).value)
        print(calo)
        if typr == 'yes':
            #print("Bot: Its a junk food... let me check your BMI")
            #a=Text(window, height=10, width=40,bg="White", fg="black",relief='flat',font=('Times',15,'bold'))
            #a.insert(END,"Bot: Its a junk food... let me check your BMI")
            #a.place(x=600,y=500)
            #Height = float(input("Bot: Enter height : "));
            #Weight = float(input("Bot: Enter Weight : "));
            Height=float(n1.get())
            
            Weight=float(n2.get())
            bmi = BMI(Height, Weight)
            if (bmi < 18.5):
                a=Text(window, height=10, width=40,bg="White", fg="black",relief='flat',font=('Times',15,'bold'))
                a.insert(END,"Bot:Healthy\ncalories-",text,format(calo))
               
                a.place(x=700,y=200)
                break
                  
            elif ( bmi >= 18.5 and bmi < 24.9):
                a=Text(window, height=10, width=40,bg="White", fg="black",relief='flat',font=('Times',15,'bold'))
                a.insert(END,"Bot: Healthy\ncalories-",c,format(calo))
               
                a.place(x=700,y=200)
                #print("Bot: Healthy")
                #print("Bot: Do cycling 5 mins.. Thats enough")
                break 
            elif ( bmi >= 24.9):
                #print("Bot: overweight")
                
                #print("Bot: Do pushups 15 count 3 reps")
                a=Text(window, height=10, width=40,bg="White", fg="black",relief='flat',font=('Times',15,'bold'))
                a.insert(END,"Bot:junk food\ncalories-",c,format(calo))
               
                a.place(x=700,y=200)
               
                break
        elif typr == 'no':
           
            a=Text(window, height=10, width=40,bg="White", fg="black",relief='flat',font=('Times',15,'bold'))
            a.insert(END,"Bot: Healthy food\nBot:calories ",text,format(calo))
            a.place(x=700,y=200)
      cv2.putText(img_to_detect,text,(45,60),cv2.FONT_HERSHEY_SIMPLEX,1.25,(255,0,0),5) 
      cv2.imshow("Detection Output", img_to_detect)
def cal():
    Height=float(n1.get())
    Weight=float(n2.get())
    test= BMI(Height, Weight)
    if (test < 18.5):
        a=Text(window, height=10, width=40,bg="White", fg="black",relief='flat',font=('Times',15,'bold'))
        a.insert(END,"Bot: No excess calories found\nBot:under weight")
        a.place(x=700,y=200)
    if (test >= 18.5 and test < 24.9):
          a=Text(window, height=10, width=40,bg="White", fg="black",relief='flat',font=('Times',15,'bold'))
          a.insert(END,"Bot:over weight\nBot: Do cycling 10 mins daily",)
          a.place(x=700,y=200)
    if(test  >= 24.9):
        a=Text(window, height=10, width=40,bg="White", fg="black",relief='flat',font=('Times',15,'bold'))
        a.insert(END,"Bot:over weight\nBot: Do cycling 10 mins daily",)
        a.place(x=700,y=200)

b2=Button(window,text='CHECK YOUR BMI',command=cal,bg='white',fg='grey37',width=30,height=1,font=('times',15,'bold'))
b2.place(x=500,y=500)


    
b1=Button(window,text='SELECT YOUR FOOD',command=open,bg='white',fg='grey37',width=30,height=1,font=('times',15,'bold'))
b1.place(x=900,y=500)

n1 = tk.StringVar()
n2 = tk.StringVar()
la=Label(window,text='Enter Your Height In Meter',bg='white',font=('times',12,'bold'))
la.place(x=200,y=250)
la1=Label(window,text='Enter Your Weight In Kg',bg='white',font=('times',12,'bold'))
la1.place(x=200,y=280)
num1= tk.Entry(window, textvariable=n1,width=30,)


num1.place(x=400,y=250)
num2= tk.Entry(window, textvariable=n2,width=30)
num2.place(x=400,y=280)




cv2.waitKey(0)
cv2.destroyAllWindows()
window.mainloop()